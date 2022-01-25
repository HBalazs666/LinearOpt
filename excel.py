from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook, load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border


#k, j, i-szerint kell indexelni!
def excel(x, xx, irange, jrange, krange):
    wb = Workbook()
    ws = wb.active
    ws.title = "MS"

    ijk = []
    ijkxx = []

    for k in range(krange):
        ijk.append([])
        for j in range(jrange):
            ijk[k].append([])
            for i in range(irange):
                ijk[k][j].append("x_{}{}{}".format(i, j, k))

    for k in range(krange):
        ijkxx.append([])
        for j in range(jrange):
            ijkxx[k].append([])
            for i in range(irange):
                ijkxx[k][j].append("xx_{}{}{}".format(i, j, k))

    for k in range(krange):
        for j in range(jrange):
            list = []
            for i in range(irange):
                list.append(ijk[k][j][i])
            ws.append(list)
        ws.append([])

# backupoké
    row = 0
    for k in range(krange):
        for j in range(jrange):
            col = irange + 5
            char = get_column_letter(irange + 5)
            for i in range(irange):
                ws.cell(row=row + 1, column=col).value = ijkxx[k][j][i]
                col += 1
            row += 1
        row += 1


    #összevetés és színezés

    yellowFill = PatternFill(start_color='00FFFF00',
                             end_color='00FFFF00',
                             fill_type='solid')

    orangeFill = PatternFill(start_color='00FF9900',
                             end_color='00FF9900',
                             fill_type='solid')

    lb = 0
    ub = 0

    for k in range(krange):
        for row in range(lb + 1, ub + jrange + 1):
            for col in range(1, irange + 1):
                char = get_column_letter(col)
                if x[k][row - 1 - ub][col - 1] == 1:
                    ws[char + str(row)].fill = yellowFill
                if xx[k][row - 1 - ub][col - 1] == 1:
                    ws[char + str(row)].fill = orangeFill
        lb += jrange + 1
        ub += jrange + 1

    lb = 0
    ub = 0

    for k in range(krange):
        for row in range(lb + 1, ub + jrange + 1):
            for col in range(1, irange + 1):
                char = get_column_letter(col + irange + 4)
                if xx[k][row - 1 - ub][col - 1] == 1:
                    ws[char + str(row)].fill = orangeFill
        lb += jrange + 1
        ub += jrange + 1


    wb.save("dsa.xlsx")
